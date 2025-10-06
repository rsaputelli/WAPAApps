
import io
import re
import pandas as pd
import streamlit as st
import altair as alt

# ============================
# Constants / Column names
# ============================
RAW_MEMBER_TYPE_COL = "Member/Non-Member - Member Type"
RAW_MEMBERSHIP_COL  = "Member/Non-Member - Membership"
ID_COL        = "Member/Non-Member - Website ID"
LAST_DUES_COL = "Member/Non-Member - Date Last Dues Transaction"
EXP_COL       = "Member/Non-Member - Date Membership Expires"
TRANSACTION_HINT_COLS = {"Dues - Amount", "Dues - Date Processed", "Dues - Date Submitted"}
HEADER_MAP = {
    "Member Type - Member Type": RAW_MEMBER_TYPE_COL,
    "Dues - Membership": RAW_MEMBERSHIP_COL,
    "Member/Non-Member - Date Membership Expires": EXP_COL,
    "Member/Non-Member - Date Last Dues Transaction": LAST_DUES_COL,
    "Member/Non-Member - Website ID": ID_COL,
    "Member/Non-Member - Email": "Member/Non-Member - Email",
    "Member/Non-Member - First Name": "Member/Non-Member - First Name",
    "Member/Non-Member - Last Name": "Member/Non-Member - Last Name",
}

# ============================
# Filename/date helpers
# ============================
def _date_prefix_today():
    import datetime as _dt
    return _dt.date.today().strftime("%Y.%m.%d")

_DATE_PATTERN = re.compile(r"^(\d{4}\.\d{2}\.\d{2})[_\-]")
def _extract_date_label(uploaded_file, fallback: str) -> str:
    name = getattr(uploaded_file, "name", "")
    m = _DATE_PATTERN.match(name)
    return m.group(1) if m else fallback

# ============================
# IO helpers
# ============================
def _normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns={c: c.strip() for c in df.columns}).rename(columns=HEADER_MAP)

def _read_any_table(uploaded_file, sheet_hint: str = None):
    """Read CSV/XLSX/XLS and try to pick a relevant sheet if Excel."""
    name = getattr(uploaded_file, "name", "").lower()
    try:
        if name.endswith(".csv"):
            try:
                return pd.read_csv(uploaded_file), None
            except UnicodeDecodeError:
                uploaded_file.seek(0)
                return pd.read_csv(uploaded_file, encoding="latin-1"), None
        else:
            xls = pd.ExcelFile(uploaded_file)
            if sheet_hint and sheet_hint in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_hint)
                return df, sheet_hint
            # Heuristic: pick sheet that contains ID or Member Type column
            for sh in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sh)
                if (ID_COL in df.columns) or (RAW_MEMBER_TYPE_COL in df.columns) or ("Member Type" in df.columns):
                    return df, sh
            sh = xls.sheet_names[0]
            df = pd.read_excel(xls, sheet_name=sh)
            return df, sh
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

# ============================
# Detection / preprocessing
# ============================
def _detect_report_type(df: pd.DataFrame) -> str:
    cols = set(df.columns)
    has_trans_cols = TRANSACTION_HINT_COLS.issubset(cols)
    has_dupes = (ID_COL in df.columns) and df[ID_COL].duplicated().any()
    return "transaction" if (has_trans_cols or has_dupes) else "member"

def _dedup_transaction_frame(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in (LAST_DUES_COL, EXP_COL):
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    is_student = df.get(RAW_MEMBER_TYPE_COL, pd.Series(dtype=object)).astype(str).str.strip().str.lower().eq("student")
    df["__is_student"] = is_student
    sort_cols, asc = ["__is_student"], [True]
    if LAST_DUES_COL in df.columns:
        sort_cols.append(LAST_DUES_COL); asc.append(False)
    if EXP_COL in df.columns:
        sort_cols.append(EXP_COL); asc.append(False)
    if ID_COL in df.columns:
        sort_cols.append(ID_COL); asc.append(True)
    df = df.sort_values(by=sort_cols, ascending=asc)
    out = df.drop_duplicates(subset=[ID_COL], keep="first").drop(columns="__is_student", errors="ignore")
    return out

def preprocess_input(df_raw: pd.DataFrame, mode: str = "auto"):
    df = _normalize_headers(df_raw)
    if RAW_MEMBER_TYPE_COL not in df.columns and "Member Type - Member Type" in df.columns:
        df = df.rename(columns={"Member Type - Member Type": RAW_MEMBER_TYPE_COL})
    if RAW_MEMBERSHIP_COL not in df.columns and "Dues - Membership" in df.columns:
        df = df.rename(columns={"Dues - Membership": RAW_MEMBERSHIP_COL})
    detected = _detect_report_type(df) if mode == "auto" else mode
    info = {"detected_mode": detected}
    if detected == "transaction":
        df_clean = _dedup_transaction_frame(df)
        info.update({"rows_in": len(df), "rows_out": len(df_clean),
                     "unique_ids_in": (df[ID_COL].nunique() if ID_COL in df.columns else None),
                     "unique_ids_out": (df_clean[ID_COL].nunique() if ID_COL in df_clean.columns else None),
                     "dedup_applied": True})
    else:
        df_clean = df
        info.update({"rows_in": len(df), "rows_out": len(df_clean),
                     "unique_ids_in": (df[ID_COL].nunique() if ID_COL in df.columns else None),
                     "unique_ids_out": (df_clean[ID_COL].nunique() if ID_COL in df_clean.columns else None),
                     "dedup_applied": False})
    for needed in (RAW_MEMBER_TYPE_COL, RAW_MEMBERSHIP_COL):
        if needed not in df_clean.columns:
            df_clean[needed] = ""
    return df_clean, info

def compute_summary_from_raw(df: pd.DataFrame):
    if RAW_MEMBER_TYPE_COL not in df.columns or RAW_MEMBERSHIP_COL not in df.columns:
        raise ValueError("Missing required columns.")
    member_type_totals = df[RAW_MEMBER_TYPE_COL].fillna("(blank)").astype(str).value_counts().sort_index()
    membership_breakdown = (
        df[[RAW_MEMBER_TYPE_COL, RAW_MEMBERSHIP_COL]].fillna("(blank)").astype(str)
        .value_counts().rename("count").reset_index()
        .sort_values([RAW_MEMBER_TYPE_COL, RAW_MEMBERSHIP_COL])
    )
    return member_type_totals, membership_breakdown

# ============================
# Legacy totals helpers (summary workbooks)
# ============================
def is_legacy_totals(df0: pd.DataFrame) -> bool:
    cols = set([c.strip() for c in df0.columns])
    return {"index","Member/Non-Member - Member Type"}.issubset(cols)

def to_totals_from_legacy(df0: pd.DataFrame) -> pd.DataFrame:
    out = df0.rename(columns={"index":"Member Type","Member/Non-Member - Member Type":"count"})[["Member Type","count"]].copy()
    out["Member Type"] = out["Member Type"].astype(str)
    out["count"] = pd.to_numeric(out["count"], errors="coerce").fillna(0).astype(int)
    return out

def to_totals_from_members(df0: pd.DataFrame) -> pd.DataFrame:
    if RAW_MEMBER_TYPE_COL not in df0.columns:
        return pd.DataFrame(columns=["Member Type","count"])
    tot = df0[RAW_MEMBER_TYPE_COL].fillna("(blank)").astype(str).value_counts().sort_index()
    return tot.rename_axis("Member Type").reset_index(name="count")

# ============================
# Legacy XLSX export (Tab1)
# ============================
def build_legacy_summary_xlsx(totals_df: pd.DataFrame, membership_breakdown: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws1 = wb.active; ws1.title = "Member Type Totals"
    legacy_totals = totals_df.rename(columns={"Member Type":"index", "count":"Member/Non-Member - Member Type"})
    for r in dataframe_to_rows(legacy_totals, index=False, header=True):
        ws1.append(r)

    chart = BarChart(); chart.type = "col"; chart.title = "Member Type Totals"
    data_ref = Reference(ws1, min_col=2, min_row=1, max_col=2, max_row=ws1.max_row)
    cats_ref = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
    chart.add_data(data_ref, titles_from_data=True); chart.set_categories(cats_ref)
    chart.x_axis.title = "Member Type"; chart.y_axis.title = "Count"
    ws1.add_chart(chart, "E2")

    ws2 = wb.create_sheet("Membership Breakdown")
    legacy_breakdown = membership_breakdown.rename(columns={"count":"Count"})
    for r in dataframe_to_rows(legacy_breakdown, index=False, header=True):
        ws2.append(r)

    import tempfile
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name); tmp.seek(0); data = tmp.read(); tmp.close()
    return data

# ============================
# Streamlit UI
# ============================
st.set_page_config(page_title="Member Analysis", layout="wide")
st.title("WAPA – Member Type Summary & Trend Builder")
st.caption(f"Run Date: {pd.Timestamp.today().date()}")

tab1, tab2, tab3 = st.tabs(["Create Monthly Summary from Raw Export", "Analyze Month-Over-Month Changes", "Build Master Workbook from Summaries"])

# ---------------- Tab 1
with tab1:
    uploaded = st.file_uploader("Choose CSV/Excel export", type=["csv","xlsx","xls"], key="single")
    if uploaded is None:
        st.info("Upload a YM CSV or Excel: either 'Today & After' (transaction-level) or 'As-of' (member-level).")
    else:
        df_raw, used_sheet = _read_any_table(uploaded)
        if used_sheet: st.caption(f"Using sheet: {used_sheet}")
        mode = st.radio("Report type", ["Auto","Transaction-level","Member-level"], index=0, horizontal=True, key="mode_single")
        mode_map = {"Auto":"auto","Transaction-level":"transaction","Member-level":"member"}
        df_clean, info = preprocess_input(df_raw, mode=mode_map[mode])

        with st.expander("Input detection & dedup summary", expanded=False):
            st.json(info); st.dataframe(df_clean.head(20))

        member_type_totals, membership_breakdown = compute_summary_from_raw(df_clean)

        st.subheader("Member Type Totals")
        totals_df = member_type_totals.rename_axis("Member Type").reset_index(name="count")
        st.dataframe(totals_df, use_container_width=True)
        st.altair_chart(alt.Chart(totals_df).mark_bar().encode(
            x=alt.X("Member Type:N", sort=None), y="count:Q", tooltip=["Member Type","count"]
        ).properties(height=260), use_container_width=True)

        st.subheader("Membership Breakdown")
        st.dataframe(membership_breakdown, use_container_width=True)

        legacy_bytes = build_legacy_summary_xlsx(totals_df, membership_breakdown)
        prefix = _date_prefix_today()
        st.download_button("Download Member_Type_Summary.xlsx (legacy format)", data=legacy_bytes,
                           file_name=f"{prefix}_Member_Type_Summary.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        csv_buf = io.StringIO(); df_clean.to_csv(csv_buf, index=False)
        st.download_button("Download cleaned member-level CSV", data=csv_buf.getvalue(),
                           file_name=f"{prefix}_member_level_clean.csv", mime="text/csv")

# ---------------- Tab 2: Month-over-Month
with tab2:
    c1, c2 = st.columns(2)
    with c1:
        f_left = st.file_uploader("Left report (e.g., Last Month)", type=["csv","xlsx","xls"], key="left")
        mode_left = st.radio("Left report type", ["Auto","Transaction-level","Member-level"], index=0, horizontal=True, key="mode_left")
    with c2:
        f_right = st.file_uploader("Right report (e.g., This Month)", type=["csv","xlsx","xls"], key="right")
        mode_right = st.radio("Right report type", ["Auto","Transaction-level","Member-level"], index=0, horizontal=True, key="mode_right")

    chart_style = st.selectbox("Chart style", ["Bars (side-by-side)", "Line (per type)"], index=0, key="chart_style_mom")

    if f_left and f_right:
        df_left_raw, used_left_sheet = _read_any_table(f_left)
        if used_left_sheet: st.caption(f"Left: using sheet '{used_left_sheet}'")
        df_right_raw, used_right_sheet = _read_any_table(f_right)
        if used_right_sheet: st.caption(f"Right: using sheet '{used_right_sheet}'")
        m = {"Auto":"auto","Transaction-level":"transaction","Member-level":"member"}
        df_left, info_left = preprocess_input(df_left_raw, mode=m[mode_left])
        df_right, info_right = preprocess_input(df_right_raw, mode=m[mode_right])

        st.markdown("#### Detection & Dedup Summaries")
        colA, colB = st.columns(2)
        with colA: st.json(info_left)
        with colB: st.json(info_right)

        has_ids_left = ID_COL in df_left.columns
        has_ids_right = ID_COL in df_right.columns

        if has_ids_left and has_ids_right:
            # Member-level compare
            left_ids  = set(df_left[ID_COL].dropna().astype(str))
            right_ids = set(df_right[ID_COL].dropna().astype(str))
            in_both   = sorted(left_ids & right_ids)
            only_left = sorted(left_ids - right_ids)
            only_right= sorted(right_ids - left_ids)

            df_both = df_left[df_left[ID_COL].astype(str).isin(in_both)].copy()
            df_only_left = df_left[df_left[ID_COL].astype(str).isin(only_left)].copy()
            df_only_right = df_right[df_right[ID_COL].astype(str).isin(only_right)].copy()

            st.markdown("### Results (Member-level)")
            st.write({"Overlap (members in both)": len(in_both), "Only in Left": len(only_left), "Only in Right": len(only_right)})
            cols_needed = [ID_COL, "Member/Non-Member - First Name", "Member/Non-Member - Last Name",
                           "Member/Non-Member - Email", RAW_MEMBER_TYPE_COL, EXP_COL, LAST_DUES_COL]
            cols_present = [c for c in cols_needed if c in df_both.columns]
            st.dataframe(df_both[cols_present].sort_values(by=ID_COL), use_container_width=True)

            c3, c4 = st.columns(2)
            with c3:
                st.markdown("#### Only in Left")
                cols_present_l = [c for c in cols_needed if c in df_only_left.columns]
                st.dataframe(df_only_left[cols_present_l].sort_values(by=ID_COL), use_container_width=True)
            with c4:
                st.markdown("#### Only in Right")
                cols_present_r = [c for c in cols_needed if c in df_only_right.columns]
                st.dataframe(df_only_right[cols_present_r].sort_values(by=ID_COL), use_container_width=True)

            # Charts from totals
            left_totals_df = to_totals_from_members(df_left)
            right_totals_df = to_totals_from_members(df_right)
            left_label = _extract_date_label(f_left, "Left")
            right_label = _extract_date_label(f_right, "Right")
            comp = pd.merge(left_totals_df.rename(columns={"count": left_label}),
                            right_totals_df.rename(columns={"count": right_label}),
                            on="Member Type", how="outer").fillna(0)
            comp["Delta (Right-Left)"] = comp[right_label] - comp[left_label]

            st.markdown("#### Member Type — Charts")
            col1, col2 = st.columns(2)
            if chart_style == "Bars (side-by-side)":
                src = comp.melt(id_vars=["Member Type"], var_name="Report", value_name="Count")
                with col1:
                    st.markdown(f"**Counts: {left_label} vs {right_label}**")
                    chart1 = alt.Chart(src).mark_bar().encode(
                        x=alt.X("Member Type:N", sort=None),
                        xOffset=alt.XOffset('Report:N'),
                        y=alt.Y("Count:Q", stack=None),
                        color="Report:N",
                        tooltip=["Member Type","Report","Count"]
                    ).properties(height=300)
                    st.altair_chart(chart1, use_container_width=True)
                with col2:
                    st.markdown("**Change (Right - Left)**")
                    chart2 = alt.Chart(comp).mark_bar().encode(
                        x=alt.X("Member Type:N", sort=None),
                        y=alt.Y("Delta (Right-Left):Q"),
                        tooltip=["Member Type","Delta (Right-Left)"]
                    ).properties(title=f"{right_label} − {left_label}", height=300)
                    st.altair_chart(chart2, use_container_width=True)
            else:
                melted = comp.melt(id_vars=["Member Type"], value_vars=[left_label, right_label], var_name="Report", value_name="Count")
                with col1:
                    st.markdown("**Counts over months**")
                    line1 = alt.Chart(melted).mark_line(point=True).encode(
                        x=alt.X("Report:N", sort=[left_label, right_label], title="Report"),
                        y="Count:Q", color="Member Type:N", tooltip=["Member Type","Report","Count"]
                    ).properties(height=300)
                    st.altair_chart(line1, use_container_width=True)
                with col2:
                    st.markdown("**Change by Member Type**")
                    chart2 = alt.Chart(comp).mark_bar().encode(
                        x=alt.X("Member Type:N", sort=None),
                        y=alt.Y("Delta (Right-Left):Q"),
                        tooltip=["Member Type","Delta (Right-Left)"]
                    ).properties(title=f"{right_label} − {left_label}", height=300)
                    st.altair_chart(chart2, use_container_width=True)

            # Excel output with charts
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.chart import BarChart, Reference
            wb = Workbook()
            ws_counts = wb.active; ws_counts.title = "Counts"
            counts_out = comp[["Member Type", left_label, right_label, "Delta (Right-Left)"]]
            for r in dataframe_to_rows(counts_out, index=False, header=True): ws_counts.append(r)
            chart_counts = BarChart(); chart_counts.type = "col"; chart_counts.title = f"{left_label} vs {right_label}"
            data_ref = Reference(ws_counts, min_col=2, min_row=1, max_col=3, max_row=ws_counts.max_row)
            cats_ref = Reference(ws_counts, min_col=1, min_row=2, max_row=ws_counts.max_row)
            chart_counts.add_data(data_ref, titles_from_data=True); chart_counts.set_categories(cats_ref)
            chart_counts.x_axis.title="Member Type"; chart_counts.y_axis.title="Count"; ws_counts.add_chart(chart_counts, "F2")
            chart_delta = BarChart(); chart_delta.type = "col"; chart_delta.title = f"Delta ({right_label} - {left_label})"
            delta_ref = Reference(ws_counts, min_col=4, min_row=1, max_col=4, max_row=ws_counts.max_row)
            chart_delta.add_data(delta_ref, titles_from_data=True); chart_delta.set_categories(cats_ref)
            chart_delta.x_axis.title="Member Type"; chart_delta.y_axis.title="Change"; ws_counts.add_chart(chart_delta, "F20")
            import tempfile; tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmp.name); tmp.seek(0); xlsx_bytes = tmp.read(); tmp.close()
            prefix = _date_prefix_today()
            st.download_button("Download comparison workbook (.xlsx)", data=xlsx_bytes,
                               file_name=f"{prefix}_Report_Comparison.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        else:
            # Summary-level compare
            left_label = _extract_date_label(f_left, "Left")
            right_label = _extract_date_label(f_right, "Right")
            totals_left = to_totals_from_legacy(df_left_raw) if is_legacy_totals(df_left_raw) else to_totals_from_members(df_left)
            totals_right = to_totals_from_legacy(df_right_raw) if is_legacy_totals(df_right_raw) else to_totals_from_members(df_right)
            all_types = sorted(set(totals_left["Member Type"]).union(set(totals_right["Member Type"])))
            df_comp = pd.DataFrame({"Member Type": all_types})
            df_comp = df_comp.merge(totals_left.rename(columns={"count":left_label}), on="Member Type", how="left")
            df_comp = df_comp.merge(totals_right.rename(columns={"count":right_label}), on="Member Type", how="left")
            df_comp[left_label] = df_comp[left_label].fillna(0).astype(int)
            df_comp[right_label] = df_comp[right_label].fillna(0).astype(int)
            df_comp["Delta (Right-Left)"] = df_comp[right_label] - df_comp[left_label]

            st.markdown("### Results (Summary-level)")
            st.dataframe(df_comp, use_container_width=True)

            st.markdown("#### Member Type — Charts")
            col1, col2 = st.columns(2)
            chart_df = df_comp.melt(id_vars=["Member Type"], value_vars=[left_label, right_label], var_name="Report", value_name="Count")
            if chart_style == "Bars (side-by-side)":
                with col1:
                    st.markdown(f"**Counts: {left_label} vs {right_label}**")
                    chart1 = alt.Chart(chart_df).mark_bar().encode(
                        x=alt.X("Member Type:N", sort=None),
                        xOffset=alt.XOffset('Report:N'),
                        y=alt.Y("Count:Q", stack=None),
                        color="Report:N", tooltip=["Member Type","Report","Count"]
                    ).properties(height=300)
                    st.altair_chart(chart1, use_container_width=True)
            else:
                with col1:
                    st.markdown("**Counts over months**")
                    line1 = alt.Chart(chart_df).mark_line(point=True).encode(
                        x=alt.X("Report:N", sort=[left_label, right_label], title="Report"),
                        y="Count:Q", color="Member Type:N", tooltip=["Member Type","Report","Count"]
                    ).properties(height=300)
                    st.altair_chart(line1, use_container_width=True)
            with col2:
                st.markdown("**Change (Right - Left)**")
                chart2 = alt.Chart(df_comp).mark_bar().encode(
                    x=alt.X("Member Type:N", sort=None),
                    y=alt.Y("Delta (Right-Left):Q"),
                    tooltip=["Member Type","Delta (Right-Left)"]
                ).properties(title=f"{right_label} − {left_label}", height=300)
                st.altair_chart(chart2, use_container_width=True)

            # Excel with charts
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.chart import BarChart, Reference
            wb = Workbook(); ws = wb.active; ws.title = "Summary Comparison"
            for r in dataframe_to_rows(df_comp, index=False, header=True): ws.append(r)
            chart_counts = BarChart(); chart_counts.type = "col"; chart_counts.title = f"{left_label} vs {right_label}"
            data_ref = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=ws.max_row)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart_counts.add_data(data_ref, titles_from_data=True); chart_counts.set_categories(cats_ref)
            chart_counts.x_axis.title="Member Type"; chart_counts.y_axis.title="Count"; ws.add_chart(chart_counts, "G2")
            chart_delta = BarChart(); chart_delta.type = "col"; chart_delta.title = f"Delta ({right_label} - {left_label})"
            delta_ref = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
            chart_delta.add_data(delta_ref, titles_from_data=True); chart_delta.set_categories(cats_ref)
            chart_delta.x_axis.title="Member Type"; chart_delta.y_axis.title="Change"; ws.add_chart(chart_delta, "G20")
            import tempfile; tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmp.name); tmp.seek(0); xlsx_bytes = tmp.read(); tmp.close()
            prefix = _date_prefix_today()
            st.download_button("Download summary comparison (.xlsx)", data=xlsx_bytes,
                               file_name=f"{prefix}_Summary_Comparison.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------- Tab 3: Longitudinal Trend Builder
with tab3:
    st.info("Upload one or more monthly summary workbooks (Member_Type_Summary.xlsx format). We'll build a running trend (line only).")
    files = st.file_uploader("Upload summary workbooks", type=["xlsx","xlsm","xls"], accept_multiple_files=True, key="multi_summaries")
    if files:
        rows = []
        for f in files:
            label = _extract_date_label(f, getattr(f, "name", "Unknown"))
            try:
                xls = pd.ExcelFile(f)
                if "Member Type Totals" not in xls.sheet_names:
                    st.error(f"{getattr(f,'name','file')}: missing required sheet 'Member Type Totals'.")
                    continue
                df_tot = pd.read_excel(xls, sheet_name="Member Type Totals")
            except Exception as e:
                st.error(f"Failed to read {getattr(f,'name','file')}: {e}")
                continue
            cols = set(df_tot.columns)
            if "index" in cols and "Member/Non-Member - Member Type" in cols:
                df_t = df_tot.rename(columns={"index":"Member Type", "Member/Non-Member - Member Type":"count"})[["Member Type","count"]].copy()
            elif "Member Type" in cols and "count" in cols:
                df_t = df_tot[["Member Type","count"]].copy()
            else:
                st.error(f"{getattr(f,'name','file')}: missing required columns in 'Member Type Totals'.")
                continue
            df_t["Date"] = label
            rows.append(df_t)

        if rows:
            df_all = pd.concat(rows, ignore_index=True)
            pivot = df_all.pivot_table(index="Date", columns="Member Type", values="count", aggfunc="sum").fillna(0).sort_index()

            st.subheader("Trend by Member Type (Line)")
            df_line = pivot.reset_index().melt(id_vars=["Date"], var_name="Member Type", value_name="Count")
            line = (
                alt.Chart(df_line)
                .mark_line(point=True)
                .encode(x="Date:N", y="Count:Q", color="Member Type:N", tooltip=["Date","Member Type","Count"])
                .properties(height=400)
            )
            st.altair_chart(line, use_container_width=True)

            # Build & download master workbook
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, Reference
            from openpyxl.utils.dataframe import dataframe_to_rows
            wb = Workbook()
            ws = wb.active; ws.title = "Monthly Totals"
            ws.append(["Date"] + list(pivot.columns))
            for idx, row in pivot.iterrows():
                ws.append([idx] + list(map(int, row.values)))
            chart = LineChart(); chart.title = "Trend by Member Type"
            data_ref = Reference(ws, min_col=2, min_row=1, max_col=1+len(pivot.columns), max_row=ws.max_row)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data_ref, titles_from_data=True); chart.set_categories(cats_ref)
            chart.y_axis.title = "Count"; chart.x_axis.title = "Month"
            from openpyxl.chart.label import DataLabelList
            for s in chart.series:
                s.dLbls = DataLabelList(); s.dLbls.showVal = True
            ws.add_chart(chart, "B{}".format(2+len(pivot.columns)))
            import tempfile
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmp.name); tmp.seek(0); xlsx_bytes = tmp.read(); tmp.close()
            prefix = _date_prefix_today()
            st.download_button("Download Master Trend Workbook (.xlsx)", data=xlsx_bytes,
                               file_name=f"{prefix}_Member_Trends.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Drop in your monthly summary workbooks to build the trend.")
